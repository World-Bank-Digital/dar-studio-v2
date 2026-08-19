"""Extract a DAMM scoring workbook into the JSON configuration the app reads.

  python3 scripts/extract-damm.py "<workbook.xlsx>" src/data/model_v1_5.json

The workbook is the methodology's source of truth; this script is the only
path from it into the application, so a new DAMM version is a re-run rather
than hand-editing 102 indicators. Output keeps the v1.3 key names the engine
already reads, and adds what v1.5 introduces (process ladder, leapfrog gap).
"""
import json
import sys
from openpyxl import load_workbook

src, out = sys.argv[1], sys.argv[2]
wb = load_workbook(src, data_only=True)


def rows(sheet, header_row):
    """Data rows below a header, skipping the leading spacer column."""
    for r in wb[sheet].iter_rows(min_row=header_row + 1, values_only=True):
        if r[1] not in (None, ""):
            yield r


def num(v):
    if v in (None, ""):
        return None
    try:
        return float(v) if isinstance(v, str) else v
    except ValueError:
        return None


# ---- anchors ------------------------------------------------------------
anchors = {}
for r in rows("Qualitative Anchors", 5):
    anchors[str(r[1])] = {f"L{i}": r[2 + i] for i in range(1, 6)}

# ---- indicators ---------------------------------------------------------
indicators = []
for r in rows("Indicators", 5):
    ind_id = str(r[1])
    cuts = [num(r[11]), num(r[12]), num(r[13]), num(r[14])]
    indicators.append({
        "id": ind_id,
        "name": r[2],
        "pillar": r[3],
        "pillar_name": r[4],
        "role": r[5],
        "source_type": r[6],
        "method": r[7],
        "direction": r[8] or "N/A",
        "gate": bool(r[9]),
        "max_age": num(r[10]),
        "cuts": None if all(c is None for c in cuts) else cuts,
        "anchors": anchors.get(ind_id, {}),
        "rubric_status": r[15],
        "calibration_note": None,
    })

# Carry forward v1.3 calibration notes for indicators v1.5 kept: the guidance
# was never withdrawn, only not re-tabulated in the new workbook.
try:
    prev = {i["id"]: i for i in json.load(open("src/data/model_v1_3.json"))["indicators"]}
    carried = 0
    for i in indicators:
        note = prev.get(i["id"], {}).get("calibration_note")
        if note:
            i["calibration_note"] = note
            carried += 1
except FileNotFoundError:
    carried = 0

# ---- config -------------------------------------------------------------
bands, coverage_gates, stage_thresholds, confidence_weights, pillars = [], {}, {}, {}, {}
leapfrog_gap = None
for r in wb["Config"].iter_rows(min_row=4, values_only=True):
    cells = [c for c in r[1:] if c not in (None, "")]
    if len(cells) < 2:
        continue
    key, val = cells[0], cells[1]
    if isinstance(key, int) and len(cells) >= 4:                     # band row
        bands.append({"level": key, "name": cells[1], "lo": cells[2], "hi": cells[3]})
    elif isinstance(key, str) and key.endswith("_min"):
        coverage_gates[key.replace("evidence_min", "evidence_adequacy_min")] = val
    elif isinstance(key, str) and key.startswith("stage"):
        stage_thresholds[key] = val
    elif key == "leapfrog_gap":
        leapfrog_gap = val
    elif key in ("High", "Medium", "Low", "Data Gap"):
        confidence_weights[key] = val
    elif isinstance(key, str) and key in ("C0", "C1", "C2", "C3", "C4", "E1", "E2", "O1"):
        role = cells[1]
        weight = cells[2] if len(cells) > 2 and isinstance(cells[2], (int, float)) else None
        entry = {"name": (cells[-1] if isinstance(cells[-1], str) else "").split(" (v1.4")[0], "role": role}
        if weight is not None:
            entry["weight"] = weight
        else:
            entry["aggregated"] = False
        pillars[key] = entry

# Pillar display names come from the indicator table, which spells them fully.
for i in indicators:
    if i["pillar"] in pillars and i["pillar_name"]:
        pillars[i["pillar"]]["name"] = str(i["pillar_name"]).split(": ", 1)[-1]

# ---- process ladder, gates, glossary ------------------------------------
process_ladder = [
    {"step": int(r[1]), "name": r[2], "executor": r[3], "output": r[4], "guidance": r[5]}
    for r in rows("Process Ladder", 5)
    if str(r[1]).strip().isdigit()
]
mandate_note = next(
    (str(r[1]) for r in rows("Process Ladder", 5)
     if not str(r[1]).strip().isdigit() and len(str(r[1])) > 60),
    None,
)
core_gates = [str(r[1]) for r in rows("Core Gates", 5)]
glossary = [{"term": r[1], "name": r[2], "text": r[3]} for r in rows("Glossary", 5)]

# Sections v1.5's workbook does not restate but the application still needs.
# roles and methodology remain true of v1.5; dar_outline is unused by the app
# (it drafts to outline.ts); ladder is carried until the 8-rung ladder is
# formally retired in favour of the 4-step process ladder.
carried_sections = {}
try:
    prev_model = json.load(open("src/data/model_v1_3.json"))
    for key in ("roles", "methodology", "dar_outline", "ladder"):
        if key in prev_model:
            carried_sections[key] = prev_model[key]
except FileNotFoundError:
    pass

model = {
    "model": "DAMM",
    "version": "1.5",
    "extracted_from": src.split("/")[-1],
    "status": "Live scoring workbook; diagnostic package workflow",
    "prohibitions": [
        "No cross-country ranking",
        "No DAMM stage used as PDO indicator, DLI, or disbursement condition",
        "No automatic financing, procurement, vendor, or technology decisions from the diagnostic",
        "No stage claimed publicly before human review by the TTL and steering committee",
    ],
    "pillars": pillars,
    "coverage_gates": coverage_gates,
    "confidence_weights": confidence_weights,
    "bands": bands,
    "stage_thresholds": stage_thresholds,
    "leapfrog_gap": leapfrog_gap,
    "assessment_year": 2026,
    "indicators": indicators,
    "process_ladder": process_ladder,
    "mandate_note": mandate_note,
    "core_gates": core_gates,
    "glossary": glossary,
    **carried_sections,
}

json.dump(model, open(out, "w"), indent=2, ensure_ascii=False)
qual = sum(1 for i in indicators if str(i["method"]).startswith("Qualitative"))
quant = sum(1 for i in indicators if i["method"] == "Quantitative threshold")
ctx = sum(1 for i in indicators if str(i["method"]).startswith("Context"))
print(f"wrote {out}")
print(f"  indicators {len(indicators)}  (quantitative {quant}, qualitative {qual}, context {ctx})")
print(f"  core gates {len(core_gates)}  pillars {len(pillars)}  bands {len(bands)}")
print(f"  anchors attached {sum(1 for i in indicators if i['anchors'])}, calibration notes carried {carried}")
print(f"  weights " + ", ".join(f"{k}={v.get('weight','ctx')}" for k, v in pillars.items()))
print(f"  leapfrog_gap {leapfrog_gap}  confidence {confidence_weights}")
