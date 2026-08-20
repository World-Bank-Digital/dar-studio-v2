"""Render the DAMM configuration the application runs on as a reference workbook.

  python3 scripts/build-damm-workbook.py [src/data/model_v1_5.json] [out.xlsx]

Every value is read from the config the engine scores against, so the workbook
cannot drift from the running model. Counts are formulas over the Indicators
sheet, not numbers computed here.
"""
import json, sys
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

SRC = sys.argv[1] if len(sys.argv) > 1 else "src/data/model_v1_5.json"
OUT = sys.argv[2] if len(sys.argv) > 2 else "exports/DAMM_v1.5_as_implemented.xlsx"
M = json.load(open(SRC))

FONT, INK, ACCENT = "Arial", "212B24", "1F5C3D"
HEAD = PatternFill("solid", fgColor="EFF3EC")
TITLE = PatternFill("solid", fgColor="1F5C3D")
GATE = PatternFill("solid", fgColor="FFF4E0")
THIN = Side(style="thin", color="DCE1D8")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
wb = Workbook()

def header(ws, row=1):
    for c in range(1, ws.max_column + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = Font(name=FONT, bold=True, size=10, color=ACCENT)
        cell.fill, cell.border = HEAD, BORDER
        cell.alignment = Alignment(vertical="center", wrap_text=True)
    ws.freeze_panes = ws.cell(row=row + 1, column=1)

def body(ws, wrap=()):
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=ws.max_column):
        for cell in row:
            cell.font = Font(name=FONT, size=10, color=INK)
            cell.border = BORDER
            cell.alignment = Alignment(vertical="top", wrap_text=cell.column_letter in wrap)

def widths(ws, spec):
    for col, w in spec.items():
        ws.column_dimensions[col].width = w

# ---------------------------------------------------------------- Read me
ws = wb.active; ws.title = "Read me"
ws["A1"] = f"{M['model']} v{M['version']} — the model this application runs"
ws["A1"].font = Font(name=FONT, bold=True, size=16, color="FFFFFF")
ws["A1"].fill = TITLE; ws.merge_cells("A1:B1"); ws.row_dimensions[1].height = 30
meta = [
    ("Model", f"{M['model']} v{M['version']}"), ("Status", M["status"]),
    ("Assessment year", M["assessment_year"]), ("Extracted from", M["extracted_from"]),
    ("Indicators", "=COUNTA(Indicators!A2:A200)"),
    ("Pillars", '=COUNTIF(Pillars!$C$2:$C$50,"Context")+COUNTIF(Pillars!$C$2:$C$50,"Capability")'
                '+COUNTIF(Pillars!$C$2:$C$50,"Ecosystem")+COUNTIF(Pillars!$C$2:$C$50,"Outcome")'),
    ("Core gates", "=COUNTA('Core gates'!A2:A50)"),
    ("Process steps", "=COUNTA('Process ladder'!A2:A50)"),
]
r = 3
for k, v in meta:
    ws.cell(row=r, column=1, value=k).font = Font(name=FONT, bold=True, size=10, color=ACCENT)
    c = ws.cell(row=r, column=2, value=v); c.font = Font(name=FONT, size=10, color=INK)
    c.alignment = Alignment(wrap_text=True, vertical="top"); r += 1
r += 1
ws.cell(row=r, column=1, value="Prohibitions wired into the model").font = Font(name=FONT, bold=True, size=11, color=ACCENT); r += 1
for i, p in enumerate(M["prohibitions"], 1):
    ws.cell(row=r, column=1, value=i).font = Font(name=FONT, size=10, color=INK)
    c = ws.cell(row=r, column=2, value=p); c.font = Font(name=FONT, size=10, color=INK)
    c.alignment = Alignment(wrap_text=True, vertical="top"); r += 1
r += 1
ws.cell(row=r, column=1, value="Sheets").font = Font(name=FONT, bold=True, size=11, color=ACCENT); r += 1
for name, desc in [
    ("Indicators", "All indicators: pillar, role, method, direction, core-gate flag, staleness limit, and the value thresholds for levels 2-5."),
    ("Qualitative anchors", "The written L1-L5 anchor text. Qualitative indicators are scored against these words, not against a number."),
    ("Pillars", "Pillars with role and weight. C0 is a context profile and is never aggregated into a score."),
    ("Core gates", "Prerequisite indicators. One at Level 1 caps the stage at 1; one unmeasured suppresses the stage entirely."),
    ("Scoring rules", "Bands, coverage gates, stage floors, the leapfrog-fragility threshold and confidence weights."),
    ("Process ladder", "The four steps: populate, score evidence quality, compile the diagnostic package, draft the roadmap."),
    ("Glossary", "CMS, EMS, OES and the rest, defined as the model defines them."),
]:
    ws.cell(row=r, column=1, value=name).font = Font(name=FONT, bold=True, size=10, color=INK)
    c = ws.cell(row=r, column=2, value=desc); c.font = Font(name=FONT, size=10, color=INK)
    c.alignment = Alignment(wrap_text=True, vertical="top"); r += 1
r += 1
ws.cell(row=r, column=1, value="Reading the thresholds").font = Font(name=FONT, bold=True, size=11, color=ACCENT); r += 1
for line in [
    "Where Direction is Higher, a value at or above the L5 threshold scores 5, at or above L4 scores 4, and so on; below the L2 threshold scores 1.",
    "Where Direction is Lower, the comparison reverses: at or below the L5 threshold scores 5.",
    "Where Direction is N/A there are no thresholds. The indicator is either a qualitative capability assessed against the anchor text, or a context profile that is never scored.",
    f"Source: {SRC.split('/')[-1]} in the application repository, extracted from {M['extracted_from']}.",
]:
    c = ws.cell(row=r, column=2, value=line); c.font = Font(name=FONT, size=10, color=INK)
    c.alignment = Alignment(wrap_text=True, vertical="top"); r += 1
widths(ws, {"A": 22, "B": 108})

# ------------------------------------------------------------- Indicators
ws = wb.create_sheet("Indicators")
ws.append(["ID", "Indicator", "Pillar", "Pillar name", "Role", "Source", "Method", "Direction",
           "Core gate", "Max age (yrs)", "L2", "L3", "L4", "L5", "Status"])
for i in M["indicators"]:
    cuts = i.get("cuts") or [None] * 4
    ws.append([i["id"], i["name"], i["pillar"], i["pillar_name"], i["role"], i["source_type"],
               i["method"], i["direction"], "YES" if i["gate"] else "", i.get("max_age"),
               *cuts, i.get("rubric_status")])
header(ws); body(ws, wrap=("B", "D", "G", "O"))
widths(ws, {"A": 7, "B": 48, "C": 7, "D": 26, "E": 12, "F": 9, "G": 24, "H": 10, "I": 10,
            "J": 12, "K": 9, "L": 9, "M": 9, "N": 9, "O": 34})
ws.auto_filter.ref = f"A1:O{ws.max_row}"
for row in range(2, ws.max_row + 1):
    if ws.cell(row=row, column=9).value == "YES":
        for c in range(1, 16):
            ws.cell(row=row, column=c).fill = GATE
        ws.cell(row=row, column=9).font = Font(name=FONT, size=10, bold=True, color="8A5A2B")

# ------------------------------------------------------ Qualitative anchors
ws = wb.create_sheet("Qualitative anchors")
ws.append(["ID", "Indicator", "Method", "Level 1", "Level 2", "Level 3", "Level 4", "Level 5"])
for i in M["indicators"]:
    a = i.get("anchors") or {}
    ws.append([i["id"], i["name"], i["method"], a.get("L1"), a.get("L2"), a.get("L3"), a.get("L4"), a.get("L5")])
header(ws); body(ws, wrap=("B", "C", "D", "E", "F", "G", "H"))
widths(ws, {"A": 7, "B": 32, "C": 20, "D": 38, "E": 38, "F": 38, "G": 38, "H": 38})
ws.auto_filter.ref = f"A1:H{ws.max_row}"

# ---------------------------------------------------------------- Pillars
ws = wb.create_sheet("Pillars")
ws.append(["Pillar", "Name", "Role", "Weight", "Aggregated", "Indicators", "Core gates"])
n = len(M["indicators"])
for pid, p in M["pillars"].items():
    row = ws.max_row + 1
    ws.append([pid, p["name"], p["role"], p.get("weight"),
               "No — context only" if p.get("aggregated") is False else "Yes",
               f'=COUNTIF(Indicators!$C$2:$C${n+1},$A{row})',
               f'=COUNTIFS(Indicators!$C$2:$C${n+1},$A{row},Indicators!$I$2:$I${n+1},"YES")'])
last = ws.max_row
ws.append(["Total", "", "", None, "", f"=SUM(F2:F{last})", f"=SUM(G2:G{last})"])
header(ws); body(ws, wrap=("B",))
for c in range(1, 8):
    ws.cell(row=last + 1, column=c).font = Font(name=FONT, bold=True, size=10, color=ACCENT)
for row in range(2, last + 1):
    ws.cell(row=row, column=4).number_format = "0%"
r2 = last + 3
ws.cell(row=r2, column=1, value="Weights sum to 1.00 within each role family").font = Font(name=FONT, bold=True, size=10, color=ACCENT)
for label, role in [("Capability (C1-C4)", "Capability"), ("Ecosystem (E1-E2)", "Ecosystem"), ("Outcome (O1)", "Outcome")]:
    r2 += 1
    ws.cell(row=r2, column=1, value=label).font = Font(name=FONT, size=10, color=INK)
    c = ws.cell(row=r2, column=4, value=f'=SUMIF($C$2:$C${last},"{role}",$D$2:$D${last})')
    c.font = Font(name=FONT, size=10, color=INK); c.number_format = "0%"
widths(ws, {"A": 34, "B": 34, "C": 13, "D": 10, "E": 20, "F": 11, "G": 11})

# ------------------------------------------------------------- Core gates
ws = wb.create_sheet("Core gates")
ws.append(["ID", "Indicator", "Pillar", "Method", "Why it gates"])
by_id = {i["id"]: i for i in M["indicators"]}
for gid in M["core_gates"]:
    g = by_id.get(gid)
    if g:
        ws.append([g["id"], g["name"], g["pillar"], g["method"],
                   "Prerequisite: at Level 1 it caps the overall stage at Stage 1; unmeasured, it suppresses the stage entirely."])
header(ws); body(ws, wrap=("B", "D", "E"))
widths(ws, {"A": 7, "B": 48, "C": 8, "D": 24, "E": 54})

# ---------------------------------------------------------- Scoring rules
ws = wb.create_sheet("Scoring rules"); r = 1
def section(title):
    global r
    ws.cell(row=r, column=1, value=title).font = Font(name=FONT, bold=True, size=11, color=ACCENT); r += 1
def table(headers, rows):
    global r
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=r, column=c, value=h)
        cell.font = Font(name=FONT, bold=True, size=10, color=ACCENT); cell.fill = HEAD; cell.border = BORDER
    r += 1
    for row in rows:
        for c, v in enumerate(row, 1):
            cell = ws.cell(row=r, column=c, value=v)
            cell.font = Font(name=FONT, size=10, color=INK); cell.border = BORDER
            cell.alignment = Alignment(wrap_text=(c >= 3), vertical="top")
        r += 1
    r += 1

section("Maturity bands — half-open and mutually exclusive")
table(["Level", "Band", "From (>=)", "To (<)"], [[b["level"], b["name"], b["lo"], b["hi"]] for b in M["bands"]])
section("Coverage gates — below these a score is suppressed rather than computed")
table(["Rule", "Threshold", "Meaning"], [
    ["pillar_min", M["coverage_gates"]["pillar_min"], "A pillar reads Not rated unless this share of its indicators carry a level."],
    ["cms_min", M["coverage_gates"]["cms_min"], "Capability coverage required before any overall stage is issued."],
    ["ems_min", M["coverage_gates"]["ems_min"], "Ecosystem coverage required before EMS is reported."],
    ["evidence_adequacy_min", M["coverage_gates"].get("evidence_adequacy_min"), "Minimum weighted evidence adequacy for a read-out to stand."],
])
section("Stage floors — the non-compensatory cascade. A stage is reached only when every floor is met.")
table(["Rule", "Floor", "Meaning"], [[k, v, f"{k.split('_')[1].upper()} floor required to be at {k.split('_')[0].replace('stage','Stage ')}."]
                                      for k, v in M["stage_thresholds"].items()])
if M.get("leapfrog_gap") is not None:
    section("Leapfrog fragility")
    table(["Rule", "Threshold", "Meaning"], [["leapfrog_gap", M["leapfrog_gap"],
        "Foundation minus transformation gap above which the assessment is flagged as leapfrog-fragile: advanced applications resting on weak foundations."]])
section("Confidence weights — applied when evidence adequacy is computed")
table(["Confidence", "Weight", "Meaning"], [[k, v, "A Data Gap is recorded and weighted zero, rather than left blank." if k == "Data Gap" else "Per-reading tag set when the reading is recorded."]
                                             for k, v in M["confidence_weights"].items()])
widths(ws, {"A": 24, "B": 12, "C": 82, "D": 10})

# --------------------------------------------------------- Process ladder
ws = wb.create_sheet("Process ladder")
ws.append(["Step", "Name", "Executor", "Output", "What happens"])
for st in M.get("process_ladder", []):
    ws.append([st["step"], st["name"], st["executor"], st["output"], st["guidance"]])
header(ws); body(ws, wrap=("B", "C", "D", "E"))
widths(ws, {"A": 7, "B": 26, "C": 28, "D": 40, "E": 78})
if M.get("mandate_note"):
    ws.append([]); ws.append(["", "Where the mandate sits", "", "", M["mandate_note"]])
    ws.cell(row=ws.max_row, column=2).font = Font(name=FONT, bold=True, size=10, color=ACCENT)
    ws.cell(row=ws.max_row, column=5).alignment = Alignment(wrap_text=True, vertical="top")
    ws.cell(row=ws.max_row, column=5).font = Font(name=FONT, size=10, color=INK)

# --------------------------------------------------------------- Glossary
ws = wb.create_sheet("Glossary")
ws.append(["Term", "Name", "Definition"])
for g in M["glossary"]:
    ws.append([g.get("term"), g.get("name"), g.get("text")])
header(ws); body(ws, wrap=("B", "C"))
widths(ws, {"A": 12, "B": 34, "C": 112})

wb.save(OUT)
print(f"wrote {OUT}")
print("  sheets:", ", ".join(s.title for s in wb.worksheets))
print(f"  {len(M['indicators'])} indicators, {len(M['core_gates'])} core gates, {len(M['pillars'])} pillars")
